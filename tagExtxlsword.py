import os
import pandas as pd
from openpyxl import load_workbook
from docx import Document
from openpyxl.styles import Font

# Load the reference list from a CSV file
reference_list = pd.read_csv('reference_list.csv')

# Create an Excel workbook to store the results
workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Tag Results'
worksheet['A1'] = 'File Name'
worksheet['B1'] = 'TAG_NAME'
worksheet['C1'] = 'Found'
worksheet['D1'] = 'Page/Sheet Number'
bold_font = Font(bold=True)
worksheet['A1'].font = bold_font
worksheet['B1'].font = bold_font
worksheet['C1'].font = bold_font
worksheet['D1'].font = bold_font

# Specify the folder containing the files (Excel and Word)
folder_path = '/path/to/files/'

# Loop through each file in the folder
for root, dirs, files in os.walk(folder_path):
    for file in files:
        file_path = os.path.join(root, file)

        # Check if the file is an Excel file
        if file.endswith(('.xlsx', '.xls')):
            excel_workbook = load_workbook(file_path, read_only=True)

            # Loop through each sheet in the Excel file
            for sheet_name in excel_workbook.sheetnames:
                sheet = excel_workbook[sheet_name]
                
                # Loop through each cell in the sheet
                for row in sheet.iter_rows(values_only=True):
                    for cell_value in row:
                        # Check if each tag in the reference list is present in the cell
                        for index, row_ref in reference_list.iterrows():
                            tag = row_ref['Tag']
                            if tag in str(cell_value):
                                worksheet.append([file, tag, 'Yes', f'Sheet: {sheet_name}'])
                            else:
                                worksheet.append([file, tag, 'No', f'Sheet: {sheet_name}'])

        # Check if the file is a Word file
        elif file.endswith('.docx'):
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                for run in paragraph.runs:
                    # Check if each tag in the reference list is present in the paragraph run
                    for index, row_ref in reference_list.iterrows():
                        tag = row_ref['Tag']
                        if tag in run.text:
                            worksheet.append([file, tag, 'Yes', 'Word Document'])
                        else:
                            worksheet.append([file, tag, 'No', 'Word Document'])

# Save the results to an Excel file with the file name
output_filename = 'tag_results.xlsx'
workbook.save(output_filename)

print(f"Results saved to {output_filename}")
